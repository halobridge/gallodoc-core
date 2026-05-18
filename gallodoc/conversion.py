"""Document → GalloMD / GalloDoc conversion (Core 2.1).

Theme: *give GalloDoc a document, get a trusted Markdown projection plus
a validated GalloDoc envelope.*

This module bridges raw document inputs (txt, md, json, csv, html, xml,
pdf, docx, xlsx, eml) and the canonical GalloDoc envelope. It is
intentionally stdlib-first — optional packages (``pypdf``, ``python-docx``,
``openpyxl``) are imported lazily and yield a clear install hint when
absent.

Public entry points:

* :func:`detect_input_type` — guess the input kind from a file path.
* :func:`build_gallomd_from_text` — produce GalloMD source from raw text.
* :func:`build_gallodoc_from_text` — produce a GalloDoc envelope.
* :func:`convert_file_to_gallomd` — read a file and return GalloMD.
* :func:`convert_file_to_gallodoc` — read a file and return a GalloDoc envelope.

The conversion never invents trust scores, signatures or hashes that
were not present in the source.
"""

from __future__ import annotations

import csv
import io
import json
import re
from dataclasses import dataclass, field
from datetime import datetime, timezone
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Iterable

from gallodoc import _MARKDOWN_SCHEMA_VERSION
from gallodoc.markdown import (
    GalloMDError,
    _scan_string_for_unsafe,
    gallomd_to_gallodoc,
)


# ---------------------------------------------------------------------------
# Errors / dataclasses
# ---------------------------------------------------------------------------


class ConversionError(RuntimeError):
    """Raised for unrecoverable conversion failures."""


@dataclass
class ConversionResult:
    """Output of a single file conversion."""

    input_path: str
    input_type: str
    title: str
    gallomd: str
    gallodoc: dict[str, Any]
    artifacts: list[dict[str, Any]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Input detection
# ---------------------------------------------------------------------------


_EXT_TO_TYPE = {
    ".txt": "txt",
    ".text": "txt",
    ".log": "txt",
    ".md": "markdown",
    ".markdown": "markdown",
    ".gmd": "gallomarkdown",
    ".gallomd": "gallomarkdown",
    ".gallomarkdown": "gallomarkdown",
    ".json": "json",
    ".csv": "csv",
    ".tsv": "csv",
    ".html": "html",
    ".htm": "html",
    ".xml": "xml",
    ".pdf": "pdf",
    ".docx": "docx",
    ".xlsx": "xlsx",
    ".eml": "eml",
    ".mbox": "eml",
}


def detect_input_type(path: str | Path) -> str:
    """Guess input type from file extension. Falls back to ``txt``."""
    ext = Path(path).suffix.lower()
    return _EXT_TO_TYPE.get(ext, "txt")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _now_iso() -> str:
    return datetime.now(tz=timezone.utc).replace(microsecond=0).isoformat()


def _slugify(text: str) -> str:
    s = re.sub(r"[^A-Za-z0-9]+", "-", (text or "").strip().lower()).strip("-")
    return s[:48] or "document"


_DATE_RE = re.compile(r"\b(?:\d{4}-\d{2}-\d{2}|\d{1,2}/\d{1,2}/\d{2,4}|[A-Z][a-z]+ \d{1,2},? \d{4})\b")
_AMOUNT_RE = re.compile(r"\$\s*[\d,]+(?:\.\d{2})?|\b\d+(?:,\d{3})*(?:\.\d{2})?\s*(?:USD|EUR|GBP)\b")
_EMAIL_RE = re.compile(r"\b[\w.+-]+@[\w-]+\.[A-Za-z]{2,}\b")
_PHONE_RE = re.compile(r"\b(?:\+?1[-.\s]?)?(?:\(\d{3}\)|\d{3})[-.\s]?\d{3}[-.\s]?\d{4}\b")


def _extract_artifacts_from_text(text: str) -> list[dict[str, Any]]:
    """Cheap, regex-based artifact extraction (dates, amounts, emails, phones)."""
    artifacts: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()

    def add(family: str, value: str) -> None:
        key = (family, value)
        if key in seen:
            return
        seen.add(key)
        artifacts.append(
            {
                "id": f"art-{family}-{len(artifacts) + 1:04d}",
                "family": family,
                "data": {"value": value},
            }
        )

    for m in _DATE_RE.finditer(text):
        add("dates", m.group())
    for m in _AMOUNT_RE.finditer(text):
        add("amounts", m.group().strip())
    for m in _EMAIL_RE.finditer(text):
        if not _scan_string_for_unsafe(m.group()):
            add("emails", m.group())
    for m in _PHONE_RE.finditer(text):
        add("phones", m.group())
    return artifacts


def _check_text_safety(text: str) -> str | None:
    """Return a short reason if ``text`` contains an obviously unsafe pattern."""
    return _scan_string_for_unsafe(text or "")


def _safe_text_or_redact(text: str) -> tuple[str, list[str]]:
    """Return a sanitized copy of ``text`` plus a list of warnings."""
    warnings: list[str] = []
    sanitized = text or ""
    reason = _scan_string_for_unsafe(sanitized)
    if reason:
        # Redact rather than block — but record the warning.
        warnings.append(f"unsafe content redacted ({reason})")
        # Replace sensitive matches.
        from gallodoc.markdown_renderer import _LINE_REDACTOR_PATTERNS, REDACTED  # local import to avoid cycle

        for pattern in _LINE_REDACTOR_PATTERNS:
            sanitized = pattern.sub(REDACTED, sanitized)
    return sanitized, warnings


# ---------------------------------------------------------------------------
# Reader implementations
# ---------------------------------------------------------------------------


def _read_txt(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")


def _read_markdown(path: Path) -> str:
    return _read_txt(path)


def _read_json(path: Path) -> tuple[str, dict[str, Any] | None]:
    raw = path.read_text(encoding="utf-8")
    try:
        data = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise ConversionError(f"invalid JSON in {path}: {exc}") from exc
    return _summarize_json(data), data


def _summarize_json(data: Any, depth: int = 0) -> str:
    if depth > 4:
        return "..."
    if isinstance(data, dict):
        lines: list[str] = []
        for k, v in list(data.items())[:40]:
            if isinstance(v, (dict, list)):
                lines.append(f"- **{k}**:")
                inner = _summarize_json(v, depth + 1)
                for line in inner.splitlines():
                    lines.append(f"    {line}")
            else:
                lines.append(f"- **{k}**: {v}")
        return "\n".join(lines)
    if isinstance(data, list):
        sample = data[:10]
        return "\n".join(f"- {_summarize_json(item, depth + 1)}" for item in sample)
    return str(data)


def _read_csv(path: Path) -> str:
    rows = []
    with path.open("r", encoding="utf-8", errors="replace", newline="") as fh:
        reader = csv.reader(fh)
        for i, row in enumerate(reader):
            rows.append(row)
            if i >= 200:
                rows.append(["..."])
                break
    if not rows:
        return ""
    header = rows[0]
    lines = ["| " + " | ".join(header) + " |"]
    lines.append("| " + " | ".join("---" for _ in header) + " |")
    for r in rows[1:25]:
        cells = list(r) + [""] * (len(header) - len(r))
        lines.append("| " + " | ".join(cells[: len(header)]) + " |")
    return "\n".join(lines)


class _HTMLTextExtractor(HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.parts: list[str] = []
        self._skip = 0
        self._headings: list[tuple[int, str]] = []
        self._cur_heading_level: int | None = None
        self._cur_heading_buf: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        if tag in ("script", "style"):
            self._skip += 1
            return
        if tag in ("h1", "h2", "h3", "h4", "h5", "h6"):
            self._cur_heading_level = int(tag[1])
            self._cur_heading_buf = []
        if tag in ("p", "br", "li", "tr", "div"):
            self.parts.append("\n")

    def handle_endtag(self, tag: str) -> None:
        if tag in ("script", "style"):
            self._skip = max(0, self._skip - 1)
            return
        if tag in ("h1", "h2", "h3", "h4", "h5", "h6") and self._cur_heading_level is not None:
            heading = "".join(self._cur_heading_buf).strip()
            if heading:
                self._headings.append((self._cur_heading_level, heading))
                self.parts.append("\n" + ("#" * self._cur_heading_level) + " " + heading + "\n")
            self._cur_heading_level = None
            self._cur_heading_buf = []

    def handle_data(self, data: str) -> None:
        if self._skip:
            return
        if self._cur_heading_level is not None:
            self._cur_heading_buf.append(data)
            return
        self.parts.append(data)


def _read_html(path: Path) -> str:
    parser = _HTMLTextExtractor()
    parser.feed(path.read_text(encoding="utf-8", errors="replace"))
    text = "".join(parser.parts)
    # Collapse runs of whitespace.
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def _read_xml(path: Path) -> str:
    raw = path.read_text(encoding="utf-8", errors="replace")
    text = re.sub(r"<[^>]+>", " ", raw)
    text = re.sub(r"\s{2,}", " ", text)
    return text.strip()


def _read_pdf(path: Path) -> tuple[str, list[str]]:
    warnings: list[str] = []
    try:
        import pypdf  # type: ignore  # noqa: PLC0415
    except ImportError:
        warnings.append(
            "pypdf not installed — PDF text extraction skipped. "
            "Install with `pip install gallodoc[pdf]`."
        )
        return f"_PDF source — install pypdf to extract text from {path.name}._", warnings
    text_parts: list[str] = []
    try:
        reader = pypdf.PdfReader(str(path))
    except Exception as exc:  # noqa: BLE001
        raise ConversionError(f"pypdf failed to open {path}: {exc}") from exc
    for i, page in enumerate(reader.pages):
        try:
            content = page.extract_text() or ""
        except Exception as exc:  # noqa: BLE001
            warnings.append(f"page {i+1} text extraction failed: {exc}")
            continue
        if content.strip():
            text_parts.append(content.strip())
    return "\n\n".join(text_parts), warnings


def _read_docx(path: Path) -> tuple[str, list[str]]:
    warnings: list[str] = []
    try:
        import docx  # type: ignore  # noqa: PLC0415
    except ImportError:
        warnings.append(
            "python-docx not installed — DOCX text extraction skipped. "
            "Install with `pip install gallodoc[docx]`."
        )
        return f"_DOCX source — install python-docx to extract text from {path.name}._", warnings
    try:
        document = docx.Document(str(path))
    except Exception as exc:  # noqa: BLE001
        raise ConversionError(f"python-docx failed to open {path}: {exc}") from exc
    parts: list[str] = []
    for para in document.paragraphs:
        text = para.text.strip()
        if not text:
            continue
        style = (para.style.name or "").lower() if para.style else ""
        if style.startswith("heading 1"):
            parts.append(f"# {text}")
        elif style.startswith("heading 2"):
            parts.append(f"## {text}")
        elif style.startswith("heading 3"):
            parts.append(f"### {text}")
        else:
            parts.append(text)
    return "\n\n".join(parts), warnings


def _read_xlsx(path: Path) -> tuple[str, list[str]]:
    warnings: list[str] = []
    try:
        from openpyxl import load_workbook  # type: ignore  # noqa: PLC0415
    except ImportError:
        warnings.append(
            "openpyxl not installed — XLSX extraction skipped. "
            "Install with `pip install openpyxl`."
        )
        return f"_XLSX source — install openpyxl to extract sheets from {path.name}._", warnings
    try:
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise ConversionError(f"openpyxl failed to open {path}: {exc}") from exc
    parts: list[str] = []
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        parts.append(f"## Sheet: {sheet_name}")
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            parts.append("_(empty)_")
            continue
        header = [str(c) if c is not None else "" for c in rows[0]]
        parts.append("| " + " | ".join(header) + " |")
        parts.append("| " + " | ".join("---" for _ in header) + " |")
        for row in rows[1:25]:
            cells = ["" if c is None else str(c) for c in row]
            cells = cells + [""] * (len(header) - len(cells))
            parts.append("| " + " | ".join(cells[: len(header)]) + " |")
    return "\n\n".join(parts), warnings


def _read_eml(path: Path) -> str:
    raw = path.read_text(encoding="utf-8", errors="replace")
    # Very small RFC-5322 friendly split: blank line separates headers/body.
    if "\n\n" in raw:
        head, _, body = raw.partition("\n\n")
    else:
        head, body = raw, ""
    headers: dict[str, str] = {}
    for line in head.splitlines():
        if ":" in line:
            k, _, v = line.partition(":")
            if k.strip().lower() in ("from", "to", "subject", "date", "cc"):
                headers[k.strip()] = v.strip()
    parts: list[str] = []
    for k in ("Subject", "From", "To", "Cc", "Date"):
        if k in headers:
            parts.append(f"- **{k}:** {headers[k]}")
    if parts:
        parts.append("")
    parts.append(body.strip())
    return "\n".join(parts)


# ---------------------------------------------------------------------------
# Synthesis: text → GalloMD source
# ---------------------------------------------------------------------------


def build_gallomd_from_text(
    text: str,
    *,
    title: str = "",
    document_type: str = "document",
    source_type: str = "text_import",
    doc_id: str = "",
    extract_artifacts: bool = False,
    structured_payload: dict[str, Any] | None = None,
    redaction_mode: str = "auto",
) -> tuple[str, list[dict[str, Any]], list[str]]:
    """Synthesize a ``.gmd`` source string from raw text.

    Returns a tuple of ``(gmd_text, extracted_artifacts, warnings)``.
    ``redaction_mode`` is one of ``"auto"``, ``"redacted"``, or ``"raw"``.
    ``"raw"`` is rejected if the text contains an obviously unsafe pattern.
    """
    if redaction_mode not in ("auto", "redacted", "raw"):
        raise ValueError(f"redaction_mode must be auto|redacted|raw, got {redaction_mode!r}")

    warnings: list[str] = []
    body = text or ""
    if redaction_mode == "raw":
        reason = _check_text_safety(body)
        if reason:
            raise ConversionError(
                f"refusing to render raw mode — unsafe content detected ({reason})"
            )
    else:
        body, redact_warnings = _safe_text_or_redact(body)
        warnings.extend(redact_warnings)

    title = title or _first_heading(body) or "Converted Document"
    if not doc_id:
        doc_id = "gmd-" + _slugify(title)
    created_at = _now_iso()

    artifacts: list[dict[str, Any]] = []
    if extract_artifacts:
        artifacts = _extract_artifacts_from_text(body)

    parts: list[str] = []
    parts.append(f"# {title}")
    parts.append("")
    parts.append("::gallodoc")
    parts.append(f"schema_version: {_MARKDOWN_SCHEMA_VERSION}")
    parts.append(f"doc_id: {doc_id}")
    parts.append(f"title: {title}")
    parts.append(f"document_type: {document_type}")
    parts.append(f"source: gallodoc_convert")
    parts.append(f"source_kind: {source_type}")
    parts.append(f"connector_slug: gallodoc_convert")
    parts.append(f"created_at: {created_at}")
    parts.append("primary_intent: conversion")
    parts.append("workflow_intent: gallodoc_convert")
    parts.append("confidence: 1.0")
    parts.append("::")
    parts.append("")
    parts.append("## Content")
    parts.append("")
    parts.append(body.strip() or f"_(empty {source_type})_")

    if structured_payload:
        parts.append("")
        parts.append("## Structured Source")
        parts.append("")
        parts.append("```json")
        try:
            parts.append(json.dumps(structured_payload, indent=2, sort_keys=True)[:4000])
        except (TypeError, ValueError):
            parts.append("(structured payload not serialisable as JSON)")
        parts.append("```")

    if artifacts:
        parts.append("")
        parts.append("## Artifacts")
        parts.append("")
        for art in artifacts:
            family = art.get("family", "generic")
            art_id = art.get("id", "")
            data = art.get("data", {})
            value = data.get("value", "")
            attrs = f"family={family}"
            if art_id:
                attrs += f" id={art_id}"
            parts.append(f"::artifact {attrs}")
            if value:
                parts.append(f"value: {value}")
            parts.append("::")
            parts.append("")

    return "\n".join(parts).rstrip() + "\n", artifacts, warnings


def _first_heading(text: str) -> str:
    for line in (text or "").splitlines():
        m = re.match(r"^#\s+(.*\S)\s*$", line)
        if m:
            return m.group(1).strip()
    return ""


# ---------------------------------------------------------------------------
# Public conversion API
# ---------------------------------------------------------------------------


def build_gallodoc_from_text(
    text: str,
    *,
    title: str = "",
    document_type: str = "document",
    source_type: str = "text_import",
    doc_id: str = "",
    extract_artifacts: bool = False,
    structured_payload: dict[str, Any] | None = None,
    redaction_mode: str = "auto",
) -> tuple[dict[str, Any], str, list[dict[str, Any]], list[str]]:
    """Synthesize a GalloDoc envelope from raw text.

    Returns ``(envelope, gallomd_source, artifacts, warnings)``.
    """
    gmd, artifacts, warnings = build_gallomd_from_text(
        text,
        title=title,
        document_type=document_type,
        source_type=source_type,
        doc_id=doc_id,
        extract_artifacts=extract_artifacts,
        structured_payload=structured_payload,
        redaction_mode=redaction_mode,
    )
    envelope = gallomd_to_gallodoc(gmd)
    return envelope, gmd, artifacts, warnings


def _read_path(path: Path) -> tuple[str, list[str], dict[str, Any] | None, str]:
    """Return ``(text, warnings, structured_payload, input_type)`` for a file."""
    input_type = detect_input_type(path)
    structured: dict[str, Any] | None = None
    warnings: list[str] = []
    if input_type == "txt":
        text = _read_txt(path)
    elif input_type in ("markdown", "gallomarkdown"):
        text = _read_markdown(path)
    elif input_type == "json":
        text, structured = _read_json(path)
    elif input_type == "csv":
        text = _read_csv(path)
    elif input_type == "html":
        text = _read_html(path)
    elif input_type == "xml":
        text = _read_xml(path)
    elif input_type == "pdf":
        text, warnings = _read_pdf(path)
    elif input_type == "docx":
        text, warnings = _read_docx(path)
    elif input_type == "xlsx":
        text, warnings = _read_xlsx(path)
    elif input_type == "eml":
        text = _read_eml(path)
    else:
        text = _read_txt(path)
    return text, warnings, structured, input_type


def convert_file_to_gallomd(
    path: str | Path,
    *,
    extract_artifacts: bool = False,
    redaction_mode: str = "auto",
) -> ConversionResult:
    """Convert a file on disk to a :class:`ConversionResult`.

    The result holds both the generated GalloMD source and the compiled
    GalloDoc envelope. Use the result fields directly or write them
    back out via the CLI.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(p)
    text, warnings, structured, input_type = _read_path(p)

    title = _first_heading(text) or p.stem.replace("_", " ").replace("-", " ").strip().title()
    doc_id = "gmd-" + _slugify(p.stem)

    if input_type == "gallomarkdown":
        # The file is already GalloMD — compile straight to JSON.
        try:
            envelope = gallomd_to_gallodoc(text)
        except GalloMDError as exc:
            raise ConversionError(f"GalloMD compile failed: {exc}") from exc
        return ConversionResult(
            input_path=str(p),
            input_type=input_type,
            title=envelope.get("identity", {}).get("title") or title,
            gallomd=text,
            gallodoc=envelope,
            artifacts=(envelope.get("extensions") or {}).get("gallomd_artifacts") or [],
            warnings=warnings,
        )

    document_type = _document_type_for(input_type)
    envelope, gmd, artifacts, build_warnings = build_gallodoc_from_text(
        text,
        title=title,
        document_type=document_type,
        source_type=input_type,
        doc_id=doc_id,
        extract_artifacts=extract_artifacts,
        structured_payload=structured,
        redaction_mode=redaction_mode,
    )
    warnings.extend(build_warnings)
    return ConversionResult(
        input_path=str(p),
        input_type=input_type,
        title=title,
        gallomd=gmd,
        gallodoc=envelope,
        artifacts=artifacts,
        warnings=warnings,
    )


def convert_file_to_gallodoc(
    path: str | Path,
    *,
    extract_artifacts: bool = False,
    redaction_mode: str = "auto",
) -> dict[str, Any]:
    """Convenience wrapper that returns just the compiled envelope."""
    return convert_file_to_gallomd(
        path,
        extract_artifacts=extract_artifacts,
        redaction_mode=redaction_mode,
    ).gallodoc


def _document_type_for(input_type: str) -> str:
    return {
        "txt": "text_document",
        "markdown": "markdown_document",
        "gallomarkdown": "gallomarkdown_document",
        "json": "structured_data",
        "csv": "tabular_data",
        "html": "html_document",
        "xml": "xml_document",
        "pdf": "pdf_document",
        "docx": "office_document",
        "xlsx": "spreadsheet",
        "eml": "email_message",
    }.get(input_type, "document")


__all__ = [
    "ConversionError",
    "ConversionResult",
    "detect_input_type",
    "build_gallomd_from_text",
    "build_gallodoc_from_text",
    "convert_file_to_gallomd",
    "convert_file_to_gallodoc",
]
